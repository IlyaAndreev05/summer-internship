"""PDF and XLSX document parsing and text chunking utilities."""

from pathlib import Path


def parse_file(path: Path) -> str:
    """Parse any supported file into its extracted text.

    Supports PDF (via pymupdf), XLSX (via openpyxl), and plain text
    files (``.txt``, ``.md``).  Returns concatenated text from all
    pages / sheets / lines.
    """
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        items = parse_pdf(path)
        return "\n\n".join(str(item["content"]) for item in items)
    if suffix in (".xlsx", ".xlsm", ".xltx", ".xltm"):
        items = parse_xlsx(path)
        return "\n\n".join(str(item["content"]) for item in items)
    if suffix in (".txt", ".md", ".csv", ".log"):
        return path.read_text(encoding="utf-8")
    raise ValueError(f"Unsupported file type: {suffix}")


def parse_pdf(path: Path) -> list[dict[str, object]]:
    """Parse a PDF file into a list of page dicts.

    Each dict has keys: ``title``, ``content``, ``page``.
    """
    import fitz  # type: ignore[import-untyped]

    results: list[dict[str, object]] = []
    doc = fitz.open(path)
    try:
        for page_num in range(len(doc)):
            page = doc[page_num]
            text = page.get_text().strip()
            if not text:
                continue
            results.append(
                {
                    "title": path.stem,
                    "content": text,
                    "page": page_num + 1,
                }
            )
    finally:
        doc.close()
    return results


def parse_xlsx(path: Path) -> list[dict[str, object]]:
    """Parse an XLSX file into a list of row dicts.

    Each dict has keys: ``title``, ``content``, ``row``.
    Content is a tab-separated representation of the row's cells.
    """
    from openpyxl import load_workbook  # type: ignore[import-untyped]

    results: list[dict[str, object]] = []
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                cells = [str(cell) if cell is not None else "" for cell in row]
                content = "\t".join(cells).strip()
                if not content:
                    continue
                results.append(
                    {
                        "title": f"{path.stem} - {sheet_name}",
                        "content": content,
                        "row": row_idx,
                    }
                )
    finally:
        wb.close()
    return results


def chunk_text(
    text: str,
    chunk_size: int = 500,
    overlap: int = 100,
) -> list[str]:
    """Split text into overlapping chunks by paragraph boundaries.

    Paragraphs (split by double newline) are grouped until they exceed
    ``chunk_size`` characters. Each chunk overlaps the previous one by
    ``overlap`` characters taken from the end of the preceding chunk.
    """
    if not text.strip():
        return []

    paragraphs = [p.strip() for p in text.split("\n\n") if p.strip()]
    chunks: list[str] = []
    current: list[str] = []
    current_len = 0

    for para in paragraphs:
        para_len = len(para)
        if current_len + para_len > chunk_size and current:
            chunk = "\n\n".join(current)
            chunks.append(chunk)
            # Build overlap: take trailing text from the finished chunk
            overlap_text = chunk[-overlap:] if len(chunk) > overlap else chunk
            current = [overlap_text]
            current_len = len(overlap_text)
        current.append(para)
        current_len += para_len + 2  # +2 for the "\n\n" separator

    if current:
        chunks.append("\n\n".join(current))

    return chunks


def ingest_file(path: Path) -> list[dict[str, object]]:
    """Parse a file and return chunked dicts.

    Each dict has keys: ``chunk_text``, ``metadata`` (a dict with
    ``source``, ``title``, ``page`` or ``row``).
    """
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        parsed = parse_pdf(path)
    elif suffix in (".xlsx", ".xlsm", ".xltx", ".xltm"):
        parsed = parse_xlsx(path)
    else:
        return []

    results: list[dict[str, object]] = []
    for item in parsed:
        content = str(item["content"])
        chunks = chunk_text(content)
        for chunk in chunks:
            meta: dict[str, str] = {
                "source": "manual",
                "title": str(item["title"]),
            }
            page = item.get("page")
            if page is not None:
                meta["page"] = str(page)
            row = item.get("row")
            if row is not None:
                meta["row"] = str(row)
            results.append({"chunk_text": chunk, "metadata": meta})
    return results


def ingest_directory(dir_path: Path) -> list[dict[str, object]]:
    """Parse all supported files in a directory and return chunked dicts."""
    suffixes = {".pdf", ".xlsx", ".xlsm", ".xltx", ".xltm"}
    results: list[dict[str, object]] = []
    for file_path in sorted(dir_path.iterdir()):
        if file_path.is_file() and file_path.suffix.lower() in suffixes:
            results.extend(ingest_file(file_path))
    return results
